# credit goes to this guy http://github.com/sh1d0wg1m3r/Metadata-Removal-Tool/blob/main/metadata_removal_tool.py#L470

import os
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

# ~(^-^)~ Pillow for images
from PIL import Image

# ~(^-^)~ PDFs
from PyPDF2 import PdfReader, PdfWriter

# ~(^-^)~ DOCX
from docx import Document

# ~(^-^)~ Audio
from mutagen.mp3 import MP3
from mutagen.id3 import ID3, ID3NoHeaderError
from mutagen.flac import FLAC

# ~(^-^)~ XLSX
from openpyxl import load_workbook

# ~(^-^)~ JPEG EXIF removal
import piexif

# ~(^-^)~ PPTX (PowerPoint 2007+); PPT is legacy
try:
    from pptx import Presentation
    CAN_HANDLE_PPTX = True
except ImportError:
    CAN_HANDLE_PPTX = False

# ~(^-^)~ ODT/ODS (OpenDocument), using odfpy if installed
try:
    from odf.opendocument import load as odf_load
    CAN_HANDLE_ODF = True
except ImportError:
    CAN_HANDLE_ODF = False

# ~(^-^)~ Logging for better production tracing
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("metadata_removal.log"),
        logging.StreamHandler()
    ]
)


################################################################
# ~(^-^)~ METADATA EXTRACTION LOGIC FOR DEMONSTRATION
################################################################

def extract_metadata_from_image(image_path):
    """
    Extract metadata from images (EXIF, etc.) for demonstration purposes
    Returns a dictionary with metadata information
    """
    metadata = {}
    try:
        ext = os.path.splitext(image_path)[1].lower()
        
        # For JPEG files, try to extract EXIF data first
        if ext in ['.jpg', '.jpeg']:
            try:
                exif_dict = piexif.load(image_path)
                exif_found = False
                
                # Convert EXIF data to readable format
                if exif_dict:
                    for category in ['0th', 'Exif', '1st', 'GPS']:
                        if category in exif_dict and exif_dict[category]:
                            for key, value in exif_dict[category].items():
                                try:
                                    tag_name = piexif.TAGS[category][key]["name"]
                                    # Convert bytes to string if needed
                                    if isinstance(value, bytes):
                                        try:
                                            clean_value = value.decode('utf-8').strip()
                                            if clean_value:  # Only add non-empty values
                                                metadata[f"EXIF_{category}_{tag_name}"] = clean_value
                                                exif_found = True
                                        except:
                                            if len(value) > 0:
                                                metadata[f"EXIF_{category}_{tag_name}"] = f"<binary: {len(value)} bytes>"
                                                exif_found = True
                                    elif isinstance(value, tuple) and len(value) == 2:
                                        # Handle rational numbers (common in EXIF)
                                        if value[1] != 0:
                                            metadata[f"EXIF_{category}_{tag_name}"] = f"{value[0]}/{value[1]} = {value[0]/value[1]:.3f}"
                                        else:
                                            metadata[f"EXIF_{category}_{tag_name}"] = str(value[0])
                                        exif_found = True
                                    elif value not in [None, '', 0]:  # Skip empty/null values
                                        metadata[f"EXIF_{category}_{tag_name}"] = str(value)
                                        exif_found = True
                                except Exception as e:
                                    pass
                
                if not exif_found:
                    logging.debug(f"No meaningful EXIF data found in {image_path}")
                
            except Exception as e:
                logging.debug(f"Could not extract EXIF from {image_path}: {e}")
        
        # Try PIL for basic image info and any additional metadata
        try:
            with Image.open(image_path) as img:
                # Get PIL info that might contain metadata
                info = img.info if hasattr(img, 'info') else {}
                for key, value in info.items():
                    if isinstance(value, (str, int, float)) and value not in [None, '', 0]:
                        # JFIF metadata is often found in PIL info
                        metadata[f"METADATA_{key}"] = str(value)
                
                # Add basic file properties (these can't be "removed" but show file characteristics)
                metadata['FILE_format'] = img.format
                metadata['FILE_mode'] = img.mode
                metadata['FILE_size_pixels'] = f"{img.size[0]}x{img.size[1]}"
        except Exception as e:
            logging.debug(f"Could not extract PIL info from {image_path}: {e}")
            
    except Exception as e:
        logging.debug(f"Error extracting image metadata from {image_path}: {e}")
    
    return metadata

def extract_metadata_from_pdf(pdf_path):
    """
    Extract metadata from PDF files for demonstration
    """
    metadata = {}
    try:
        reader = PdfReader(pdf_path)
        if reader.metadata:
            for key, value in reader.metadata.items():
                # Clean up the key name (remove leading slash)
                clean_key = key.lstrip('/')
                metadata[clean_key] = str(value) if value else ''
        
        # Add page count
        metadata['page_count'] = str(len(reader.pages))
                
    except Exception as e:
        logging.debug(f"Error extracting PDF metadata from {pdf_path}: {e}")
    
    return metadata

def extract_metadata_from_docx(docx_path):
    """
    Extract metadata from DOCX files for demonstration
    """
    metadata = {}
    try:
        doc = Document(docx_path)
        if doc.core_properties:
            props = doc.core_properties
            metadata['title'] = props.title or ''
            metadata['author'] = props.author or ''
            metadata['subject'] = props.subject or ''
            metadata['created'] = str(props.created) if props.created else ''
            metadata['modified'] = str(props.modified) if props.modified else ''
            metadata['last_modified_by'] = props.last_modified_by or ''
            metadata['revision'] = str(props.revision) if props.revision else ''
            metadata['category'] = props.category or ''
            metadata['comments'] = props.comments or ''
            
    except Exception as e:
        logging.debug(f"Error extracting DOCX metadata from {docx_path}: {e}")
    
    return metadata

def extract_metadata_from_mp3(mp3_path):
    """
    Extract metadata from MP3 files for demonstration
    """
    metadata = {}
    try:
        audio = MP3(mp3_path)
        if audio.tags:
            for key, value in audio.tags.items():
                # Convert to string and clean up
                clean_key = str(key).replace(':', '_')
                if isinstance(value, list) and value:
                    metadata[clean_key] = str(value[0])
                else:
                    metadata[clean_key] = str(value)
        
        # Add duration and bitrate
        if hasattr(audio, 'info'):
            metadata['duration'] = f"{audio.info.length:.2f} seconds" if audio.info.length else 'Unknown'
            metadata['bitrate'] = f"{audio.info.bitrate} bps" if audio.info.bitrate else 'Unknown'
            
    except Exception as e:
        logging.debug(f"Error extracting MP3 metadata from {mp3_path}: {e}")
    
    return metadata

def extract_metadata_from_xlsx(xlsx_path):
    """
    Extract metadata from XLSX files for demonstration
    """
    metadata = {}
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        if wb.properties:
            props = wb.properties
            metadata['title'] = props.title or ''
            metadata['creator'] = props.creator or ''
            metadata['subject'] = props.subject or ''
            metadata['created'] = str(props.created) if props.created else ''
            metadata['modified'] = str(props.modified) if props.modified else ''
            metadata['lastModifiedBy'] = props.lastModifiedBy or ''
            metadata['category'] = props.category or ''
            metadata['description'] = props.description or ''
        
        # Add worksheet count
        metadata['worksheet_count'] = str(len(wb.worksheets))
        wb.close()
        
    except Exception as e:
        logging.debug(f"Error extracting XLSX metadata from {xlsx_path}: {e}")
    
    return metadata

def extract_metadata(file_path):
    """
    Extract metadata from a file based on its extension for demonstration purposes
    Returns a dictionary with metadata information
    """
    metadata = {}
    try:
        file_extension = os.path.splitext(file_path)[1].lower()
        
        # Images
        if file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
            metadata = extract_metadata_from_image(file_path)
        
        # PDFs
        elif file_extension == '.pdf':
            metadata = extract_metadata_from_pdf(file_path)
        
        # DOCX
        elif file_extension == '.docx':
            metadata = extract_metadata_from_docx(file_path)
        
        # MP3
        elif file_extension == '.mp3':
            metadata = extract_metadata_from_mp3(file_path)
        
        # XLSX
        elif file_extension == '.xlsx':
            metadata = extract_metadata_from_xlsx(file_path)
        
        # Add file statistics
        if os.path.exists(file_path):
            from datetime import datetime
            stat = os.stat(file_path)
            metadata['file_size_bytes'] = str(stat.st_size)
            metadata['modified_time'] = str(datetime.fromtimestamp(stat.st_mtime))
        
    except Exception as e:
        logging.debug(f"Error extracting metadata from {file_path}: {e}")
    
    return metadata

def demo_metadata_before_after(file_path, downloads_dir=None, watermark_text="VALIDATED", original_filename=None):
    """
    Demonstration function to show metadata before and after removal
    Creates a copy of the file to preserve the original and make a clear before/after
    Saves the cleaned file for download if downloads_dir is provided
    Returns: {'before': dict, 'after': dict, 'removed_count': int, 'success': bool, 'download_id': str, 'original_filename': str}
    """
    result = {
        'before': {},
        'after': {},
        'removed_count': 0,
        'success': False,
        'details': '',
        'download_id': None,
        'original_filename': None,
        'cleaned_file_path': None
    }
    
    try:
        # Store original filename for download (use provided original_filename if available)
        result['original_filename'] = original_filename if original_filename else os.path.basename(file_path)
        
        # Extract metadata before (from original file)
        result['before'] = extract_metadata(file_path)
        logging.debug(f"Before metadata extraction: {len(result['before'])} fields found")
        
        # Create a working copy for metadata removal (preserve original)
        import shutil
        import tempfile
        import uuid
        
        # Create a temporary copy to work on
        temp_dir = tempfile.mkdtemp()
        temp_file = os.path.join(temp_dir, f"temp_{os.path.basename(file_path)}")
        shutil.copy2(file_path, temp_file)
        
        # Remove metadata from the copy
        removal_success = remove_metadata(temp_file)
        logging.debug(f"Metadata removal result: {removal_success}")
        
        # Extract metadata after (from processed copy)
        result['after'] = extract_metadata(temp_file)
        logging.debug(f"After metadata extraction: {len(result['after'])} fields found")
        
        # Count actually meaningful metadata (exclude basic file properties)
        before_metadata = {k: v for k, v in result['before'].items() 
                          if not k.startswith('FILE_') and v not in [None, '', '0']}
        after_metadata = {k: v for k, v in result['after'].items() 
                         if not k.startswith('FILE_') and v not in [None, '', '0']}
        
        result['removed_count'] = len(before_metadata) - len(after_metadata)
        result['success'] = removal_success
        
        # Save cleaned file for download if downloads directory is provided
        if downloads_dir and removal_success:
            try:
                # Ensure downloads directory exists
                os.makedirs(downloads_dir, exist_ok=True)
                
                # Generate unique download ID
                download_id = str(uuid.uuid4())
                
                # Create cleaned filename with download ID
                base_name, ext = os.path.splitext(result['original_filename'])
                cleaned_filename = f"{base_name}_cleaned_{download_id}{ext}"
                cleaned_file_path = os.path.join(downloads_dir, cleaned_filename)
                
                # Copy cleaned file to downloads directory
                shutil.copy2(temp_file, cleaned_file_path)
                
                # Apply watermark to the cleaned file if it's an image
                try:
                    # Import watermarking functionality
                    try:
                        from .watermaker import apply_red_watermark, is_image_file
                        watermark_available = True
                    except ImportError:
                        try:
                            from watermaker import apply_red_watermark, is_image_file
                            watermark_available = True
                        except ImportError:
                            watermark_available = False
                    
                    if watermark_available:
                        # Verify the cleaned file exists and is not empty
                        if os.path.exists(cleaned_file_path) and os.path.getsize(cleaned_file_path) > 0:
                            if is_image_file(cleaned_file_path):
                                watermark_result = apply_red_watermark(
                                    cleaned_file_path, 
                                    watermark_text=watermark_text,
                                    save_to_path=cleaned_file_path  # Overwrite the cleaned file
                                )
                                if watermark_result['success']:
                                    logging.info(f"Red watermark applied to cleaned file: {cleaned_file_path}")
                                    result['details'] += " + watermark applied"
                                else:
                                    error_msg = watermark_result.get('error', 'Unknown watermarking error')
                                    logging.warning(f"Watermark failed on cleaned file: {error_msg}")
                                    result['details'] += f" (watermark failed: {error_msg})"
                            else:
                                logging.info("Cleaned file is not an image, skipping watermark")
                        else:
                            logging.error(f"Cleaned file does not exist or is empty: {cleaned_file_path}")
                    else:
                        logging.info("Watermarking not available for cleaned file")
                            
                except Exception as watermark_error:
                    error_msg = str(watermark_error) if str(watermark_error) else "Unknown watermarking exception"
                    logging.error(f"Error applying watermark to cleaned file: {error_msg}")
                    # Don't fail the operation if watermarking fails
                    result['details'] += f" (watermark error: {error_msg})"
                
                result['download_id'] = download_id
                result['cleaned_file_path'] = cleaned_file_path
                
                logging.info(f"Cleaned file saved for download: {cleaned_file_path}")
                
            except Exception as download_error:
                logging.error(f"Failed to save cleaned file for download: {download_error}")
                # Don't fail the entire operation if download preparation fails
        
        # Add details about what was processed
        if len(before_metadata) == 0:
            result['details'] = 'No removable metadata found in original file'
        elif result['removed_count'] > 0:
            result['details'] = f'Successfully removed {result["removed_count"]} metadata fields'
        else:
            result['details'] = 'Metadata removal attempted but no fields were removed'
            
        # Clean up temp file
        try:
            os.unlink(temp_file)
            os.rmdir(temp_dir)
        except:
            pass
        
    except Exception as e:
        logging.error(f"Error in demo metadata before/after for {file_path}: {e}")
        result['success'] = False
        result['details'] = f'Error: {str(e)}'
    
    return result

################################################################
# ~(^-^)~ METADATA REMOVAL LOGIC FOR VARIOUS FILE FORMATS
################################################################

def remove_metadata_from_zip(zip_path):
    """ 
    Unzip -> Remove metadata from each entry -> Re-zip.
    (｡•̀ᴗ-)✧
    """
    temp_dir = "temp_zip_extract"
    try:
        logging.info(f"Processing ZIP: {zip_path}")
        os.makedirs(temp_dir, exist_ok=True)
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # ~(^-^)~ Clean each file inside temp_dir
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root, file)
                remove_metadata(file_path)

        # ~(^-^)~ Repackage as a new, cleaned ZIP
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arc_name = os.path.relpath(file_path, temp_dir)
                    zip_ref.write(file_path, arcname=arc_name)

        # ~(^-^)~ Cleanup
        for root, dirs, files in os.walk(temp_dir, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.rmdir(temp_dir)

        return True
    except Exception as e:
        logging.exception(f"Error processing ZIP {zip_path}: {e}")
        return False


def remove_exif_jpeg(file_path):
    """ 
    Strip EXIF from JPEG without re-encoding 
    ヾ(⌐■_■)ノ♪
    """
    try:
        logging.info(f"Stripping EXIF from JPEG: {file_path}")
        piexif.remove(file_path)
        # Verify removal
        exif_dict = piexif.load(file_path)
        if any(exif_dict[tag] for tag in ["0th", "Exif", "GPS", "1st"]):
            logging.info(f"EXIF still present after piexif.remove. Re-encoding: {file_path}")
            _reencode_jpeg(file_path)
        else:
            logging.info(f"EXIF successfully removed: {file_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing JPEG EXIF {file_path}: {e}")
        # Fallback to re-encode if piexif fails
        try:
            _reencode_jpeg(file_path)
            return True
        except Exception as ex:
            logging.exception(f"Failed to re-encode JPEG: {file_path}, error: {ex}")
            return False


def _reencode_jpeg(file_path):
    """
    Re-encode the JPEG with Pillow. This definitely strips EXIF,
    but may alter quality/size if not configured carefully.
    """
    try:
        with Image.open(file_path) as img:
            # Convert to RGB (some images might be in different modes)
            rgb_img = img.convert("RGB")
            # Save, forcing 'exif' to be blank
            rgb_img.save(file_path, format='JPEG', exif=b'')
        logging.info(f"Re-encoded JPEG to remove residual EXIF: {file_path}")
    except Exception as e:
        logging.exception(f"Failed to re-encode JPEG: {file_path}, error: {e}")
        raise


def remove_metadata_from_jpeg(file_path):
    """
    Attempt to remove EXIF from JPEG using piexif (no re-encoding).
    If EXIF still remains, fallback to re-encoding with Pillow to ensure
    all metadata is stripped. This may alter quality/file size slightly.
    """
    try:
        remove_exif_jpeg(file_path)
        return True
    except Exception as e:
        logging.error(f"Failed to remove metadata from JPEG {file_path}: {e}")
        return False


def remove_metadata_from_image(image_path):
    """
    Images beyond JPEG (PNG, GIF, BMP, TIFF):
    Re-encode with Pillow to drop metadata.
    (ﾉ◕ヮ◕)ﾉ*:･ﾟ✧
    """
    ext = os.path.splitext(image_path)[1].lower()
    if ext in [".jpg", ".jpeg"]:
        return remove_metadata_from_jpeg(image_path)
    try:
        logging.info(f"Processing image: {image_path}")
        with Image.open(image_path) as img:
            data = list(img.getdata())
            clean_img = Image.new(img.mode, img.size)
            clean_img.putdata(data)
            # Keep the original format if possible
            clean_img.save(image_path, format=img.format)
        logging.info(f"Metadata removed from image: {image_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing image {image_path}: {e}")
        return False


def remove_metadata_from_pdf(pdf_path):
    """
    PDF: Read pages, rewrite them, omit doc info. 
    (／・ω・)／
    """
    try:
        logging.info(f"Processing PDF: {pdf_path}")
        reader = PdfReader(pdf_path)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        with open(pdf_path, 'wb') as out_pdf:
            writer.write(out_pdf)
        logging.info(f"Metadata removed from PDF: {pdf_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing PDF {pdf_path}: {e}")
        return False


def remove_metadata_from_docx(docx_path):
    """
    DOCX: Clear core properties with python-docx.
    (ﾉ´ヮ´)ﾉ*:･ﾟ✧
    """
    try:
        logging.info(f"Processing DOCX: {docx_path}")
        doc = Document(docx_path)
        metadata_fields = [
            'author', 'comments', 'category', 'content_status',
            'identifier', 'keywords', 'language', 'last_modified_by',
            'last_printed', 'revision', 'subject', 'title', 'version'
        ]
        for prop in metadata_fields:
            try:
                setattr(doc.core_properties, prop, "")
            except ValueError:
                pass
        doc.settings.odd_and_even_pages_header_footer = False
        doc.save(docx_path)
        logging.info(f"Metadata removed from DOCX: {docx_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing DOCX {docx_path}: {e}")
        return False


def remove_metadata_from_pptx(pptx_path):
    """
    PPTX: Clear core properties with python-pptx.
    ヾ(*ΦωΦ)ツ
    """
    if not CAN_HANDLE_PPTX:
        logging.warning("python-pptx not installed; cannot process PPTX.")
        return False
    try:
        logging.info(f"Processing PPTX: {pptx_path}")
        ppt = Presentation(pptx_path)
        props = ppt.core_properties
        # ~(^-^)~ Clear 'em all
        props.author = ""
        props.category = ""
        props.comments = ""
        props.content_status = ""
        props.created = None
        props.identifier = ""
        props.keywords = ""
        props.last_modified_by = ""
        props.last_printed = None
        props.modified = None
        props.revision = ""
        props.subject = ""
        props.title = ""
        ppt.save(pptx_path)
        logging.info(f"Metadata removed from PPTX: {pptx_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing PPTX {pptx_path}: {e}")
        return False


def remove_metadata_from_ppt(ppt_path):
    """
    PPT (legacy):
    Typically convert .ppt -> .pptx, then do the PPTX route.
    (╯°□°)╯︵ ┻━┻
    """
    try:
        logging.warning(f".ppt is legacy binary; recommended to convert {ppt_path} to .pptx first.")
        return False
    except Exception as e:
        logging.exception(f"Error processing PPT {ppt_path}: {e}")
        return False


def remove_metadata_from_odt(odt_path):
    """
    ODT (OpenDocument Text): Use odfpy to remove <office:meta>.
    (ˆ-ˆ)و♪
    """
    if not CAN_HANDLE_ODF:
        logging.warning("odfpy not installed; cannot process ODT.")
        return False
    try:
        logging.info(f"Processing ODT: {odt_path}")
        doc = odf_load(odt_path)
        meta = doc.meta
        # ~(^-^)~ Remove all metadata children
        for child in list(meta.childNodes):
            meta.removeChild(child)
        doc.save(odt_path)
        logging.info(f"Metadata removed from ODT: {odt_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing ODT {odt_path}: {e}")
        return False


def remove_metadata_from_ods(ods_path):
    """
    ODS (OpenDocument Spreadsheet): Also via odfpy.
    (☞ﾟヮﾟ)☞
    """
    if not CAN_HANDLE_ODF:
        logging.warning("odfpy not installed; cannot process ODS.")
        return False
    try:
        logging.info(f"Processing ODS: {ods_path}")
        spreadsheet = odf_load(ods_path)
        meta = spreadsheet.meta
        for child in list(meta.childNodes):
            meta.removeChild(child)
        spreadsheet.save(ods_path)
        logging.info(f"Metadata removed from ODS: {ods_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing ODS {ods_path}: {e}")
        return False


def remove_metadata_from_epub(epub_path):
    """
    EPUB: Search for .opf files, remove <metadata> content.
    ヾ(〃^∇^)ﾉ
    """
    temp_dir = "temp_epub_extract"
    try:
        logging.info(f"Processing EPUB: {epub_path}")
        os.makedirs(temp_dir, exist_ok=True)
        with zipfile.ZipFile(epub_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)

        # ~(^-^)~ Look for .opf
        opf_files = []
        for root, dirs, files in os.walk(temp_dir):
            for file in files:
                if file.lower().endswith('.opf'):
                    opf_files.append(os.path.join(root, file))

        # ~(^-^)~ Strip known metadata tags
        import re
        for opf_file in opf_files:
            try:
                with open(opf_file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                # Remove <metadata> sections in a naive way
                content = re.sub(r'<metadata[^>]*>.*?</metadata>', '<metadata></metadata>', content, flags=re.DOTALL)
                with open(opf_file, 'w', encoding='utf-8') as f:
                    f.write(content)
                logging.info(f"Metadata stripped from OPF file: {opf_file}")
            except Exception as ex:
                logging.error(f"Could not strip metadata from {opf_file}: {ex}")

        # ~(^-^)~ Re-zip
        with zipfile.ZipFile(epub_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for root, dirs, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    zip_ref.write(file_path, arcname)
        logging.info(f"Re-zipped EPUB after metadata removal: {epub_path}")

        # ~(^-^)~ Cleanup
        for root, dirs, files in os.walk(temp_dir, topdown=False):
            for name in files:
                os.remove(os.path.join(root, name))
            for name in dirs:
                os.rmdir(os.path.join(root, name))
        os.rmdir(temp_dir)

        return True
    except Exception as e:
        logging.exception(f"Error processing EPUB {epub_path}: {e}")
        return False


def remove_metadata_from_rtf(rtf_path):
    r"""
    RTF: Naive \info removal with regex.
    (ﾉ´ヮ´)ﾉ*:･ﾟ✧
    """
    try:
        logging.info(f"Processing RTF: {rtf_path}")
        with open(rtf_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        import re
        # Remove {\info ...} blocks
        content = re.sub(r'{\\info[^}]*}', '', content, flags=re.DOTALL)
        with open(rtf_path, 'w', encoding='utf-8') as f:
            f.write(content)
        logging.info(f"Metadata removed from RTF: {rtf_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing RTF {rtf_path}: {e}")
        return False


def remove_metadata_from_mp3(mp3_path):
    """ 
    MP3: Delete ID3 tags with mutagen.
    (＾▽＾) 
    """
    try:
        logging.info(f"Processing MP3: {mp3_path}")
        audio = MP3(mp3_path, ID3=ID3)
        audio.delete()
        audio.save(mp3_path)
        logging.info(f"Metadata removed from MP3: {mp3_path}")
        return True
    except ID3NoHeaderError:
        try:
            audio = MP3(mp3_path)
            audio.save(mp3_path)
            logging.info(f"No ID3 header found, but file saved: {mp3_path}")
            return True
        except Exception as e:
            logging.exception(f"Error saving MP3 {mp3_path} without ID3 header: {e}")
            return False
    except Exception as e:
        logging.exception(f"Error processing MP3 {mp3_path}: {e}")
        return False


def remove_metadata_from_flac(flac_path):
    """ 
    FLAC: Remove tags with mutagen.
    (｡•̀ᴗ-)✧
    """
    try:
        logging.info(f"Processing FLAC: {flac_path}")
        audio = FLAC(flac_path)
        audio.delete()
        audio.save()
        logging.info(f"Metadata removed from FLAC: {flac_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing FLAC {flac_path}: {e}")
        return False


def remove_metadata_from_xlsx(xlsx_path):
    """
    XLSX: Clear workbook properties with openpyxl.
    (∿°○°)∿
    """
    try:
        logging.info(f"Processing XLSX: {xlsx_path}")
        workbook = load_workbook(filename=xlsx_path)
        metadata_fields = [
            'creator', 'title', 'subject', 'description',
            'keywords', 'category', 'comments', 'last_modified_by',
            'company', 'manager'
        ]
        for prop in metadata_fields:
            try:
                setattr(workbook.properties, prop, "")
            except ValueError:
                pass
        workbook.save(xlsx_path)
        logging.info(f"Metadata removed from XLSX: {xlsx_path}")
        return True
    except Exception as e:
        logging.exception(f"Error processing XLSX {xlsx_path}: {e}")
        return False


################################################################
# ~(^-^)~ MASTER SWITCH: Determine file type + call appropriate remover
################################################################

def remove_metadata(file_path):
    """ 
    Decide how to remove metadata based on file extension. 
    (ﾉ◕ヮ◕)ﾉ*:･ﾟ✧
    """
    file_extension = os.path.splitext(file_path)[1].lower()

    # ~(^-^)~ Images
    if file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
        return remove_metadata_from_image(file_path)

    # ~(^-^)~ PDFs
    elif file_extension == '.pdf':
        return remove_metadata_from_pdf(file_path)

    # ~(^-^)~ DOCX
    elif file_extension == '.docx':
        return remove_metadata_from_docx(file_path)

    # ~(^-^)~ MP3 / FLAC
    elif file_extension == '.mp3':
        return remove_metadata_from_mp3(file_path)
    elif file_extension == '.flac':
        return remove_metadata_from_flac(file_path)

    # ~(^-^)~ XLSX
    elif file_extension == '.xlsx':
        return remove_metadata_from_xlsx(file_path)

    # ~(^-^)~ ZIP
    elif file_extension == '.zip':
        return remove_metadata_from_zip(file_path)

    # ~(^-^)~ PPTX / PPT
    elif file_extension == '.pptx':
        return remove_metadata_from_pptx(file_path)
    elif file_extension == '.ppt':
        return remove_metadata_from_ppt(file_path)

    # ~(^-^)~ ODT / ODS
    elif file_extension == '.odt':
        return remove_metadata_from_odt(file_path)
    elif file_extension == '.ods':
        return remove_metadata_from_ods(file_path)

    # ~(^-^)~ EPUB
    elif file_extension == '.epub':
        return remove_metadata_from_epub(file_path)

    # ~(^-^)~ RTF
    elif file_extension == '.rtf':
        return remove_metadata_from_rtf(file_path)

    else:
        logging.warning(f"Unsupported file type {file_extension} for {file_path}.")
        return False


################################################################
# ~(^-^)~ GUI / APPLICATION LOGIC
################################################################




if __name__ == "__main__":
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "FOX621_SAMPLE1.png")
    
    if os.path.exists(file_path):
        remove_metadata(file_path)
    else:
        print(f"File not found: {file_path}")